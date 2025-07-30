from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404
from .models import *
from django.db.models import Q
from django.template.loader import render_to_string
import requests
from django.http import JsonResponse, Http404
from .sync import sincronizar_empleados
from .sync import sincronizar_todas_sucursales
from .sync_marcaje import sincronizar_marcajes
from django.core import serializers
from django.views.decorators.http import require_POST
from datetime import datetime, timedelta, date
from django.utils import timezone
from .depurar_marcajes import depurar_marcajes
from .forms import *
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from datetime import date, datetime
from django.core.mail import send_mail
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse, reverse_lazy
from django.db.models.functions import Upper, Trim
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from django.db.models import ProtectedError
from collections import defaultdict
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
import traceback
from django.conf import settings


def grupo_requerido(nombre_grupo):
    def check(user):
        return user.groups.filter(name=nombre_grupo).exists()
    def decorator(view_func):
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated or not check(request.user):
                raise Http404("P√°gina no encontrada")
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

@login_required
@grupo_requerido('rrhh')
def vacaciones_proxy(request):
    empleado = request.GET.get("empleado")
    if not empleado:
        return JsonResponse({"error": "Falta el par√°metro 'empleado'"}, status=400)

    # Aqu√≠ se construye la URL completa con el par√°metro ya insertado
    target_url = f"http://192.168.11.12:8000/planilla/webservice/vacaciones/disponibles/?empleado={empleado}"

    headers = {
        "X-API-Key": settings.API_KEY_VACACIONES,
        "X-Requested-With": "XMLHttpRequest"
    }

    try:
        response = requests.get(
            target_url,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return JsonResponse(response.json(), safe=False)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@grupo_requerido('rrhh')
def empleados_proxy(request):
    target_url = "http://192.168.11.12:8000/planilla/webservice/empleados/"
    
    headers = {
        'X-Requested-With': 'XMLHttpRequest',
        'Accept': 'application/json',
    }
    
    try:
        response = requests.get(
            target_url,
            headers=headers,
            params={'sucursal': 4},
            timeout=10
        )
        response.raise_for_status()
        return JsonResponse(response.json())
    
    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )
# @csrf_exempt  
def asistencias_api(request):
    # if request.method == 'GET':
    #     fecha_str = request.GET.get('fecha')
        
    #     if fecha_str:
    #         try:
    #             fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    #         except ValueError:
    #             return JsonResponse({'error': 'Formato de fecha inv√°lido'}, status=400)
    #     else:
    #         fecha = timezone.now().date()
            
    target_url = f"http://192.168.11.12:8003/api/asistencias/?fecha=2025-07-28"  #?fecha={fecha.strftime('%Y-%m-%d')}"
    
    headers = {
        "X-API-Key": settings.API_KEY_ASISTENCIA  # o "x-api-key": "TU_API_KEY"
    }
    
    try:
        response = requests.get(
            target_url,
            headers=headers,
            timeout=10
        )
        response.raise_for_status()
        return JsonResponse(response.json())
    
    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )
    
@login_required
@grupo_requerido('rrhh')
def sync_empleados_view(request):
    resultados = sincronizar_todas_sucursales()
    
    total_creados = sum(r.get('creados', 0) for r in resultados if r.get('status') == 'success')
    total_actualizados = sum(r.get('actualizados', 0) for r in resultados if r.get('status') == 'success')

    empleados = Empleado.objects.filter(activo=True)
    empleados_json = serializers.serialize('json', empleados)

    return JsonResponse({
        'status': 'success',
        'creados': total_creados,
        'actualizados': total_actualizados,
        'sincronizaciones': resultados,
        'empleados': empleados_json,
    })


@login_required
@grupo_requerido('rrhh')
@require_POST
def sync_marcaje_view(request):
    try:
        fecha_str = request.GET.get('fecha')
        if not fecha_str:
            fecha_str = timezone.now().strftime('%Y-%m-%d')
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()

        resultado = sincronizar_marcajes(fecha=fecha)

        if 'error' in resultado:
            return JsonResponse({
                'status': 'error',
                'message': resultado['error'],
                'marcajes': []
            }, status=400)

        depurar_marcajes(fecha)

        marcajes = Marcaje.objects.all().order_by('-fecha_hora')

        return JsonResponse({
            'status': 'success',
            'message': f'Sincronizaci√≥n completada para {fecha}',
            'creados': resultado.get('creados', 0),
            'actualizados': resultado.get('actualizados', 0),
            'errores': resultado.get('errores', 0),
            'marcajes': serializers.serialize('json', marcajes)
        })

    except Exception as e:
        tb = traceback.format_exc()
        print(tb)  # Ver detalle en consola
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'traceback': tb,
            'marcajes': []
        }, status=500)


@login_required
@grupo_requerido('rrhh')
def marcar(request):
    departamento = request.GET.get('departamento')
    empleados = Empleado.objects.filter(activo=True)

    if departamento:
        empleados = empleados.filter(departamento=departamento)
    
    departamentos = Empleado.objects.order_by('departamento'
                      ).values_list('departamento', flat=True
                      ).distinct()

    context = {
        'empleados': empleados,
        'departamentos': departamentos,
        
        'departamento_seleccionado': departamento,
    }
    return render(request, 'empleados.html', context)

@login_required
@grupo_requerido('rrhh')
def lista_registros(request):
    departamento = request.GET.get('departamento')
    registros = Marcaje.objects.select_related('empleado').all() 
    if departamento:
        registros = registros.filter(empleado__departamento=departamento)
    
    registros = registros.order_by('-fecha_hora')
    
    departamentos = Empleado.objects.values_list('departamento', flat=True).distinct().order_by()

    paginator = Paginator(registros, 100)
    page_number = request.GET.get('page')
    registros = paginator.get_page(page_number)

    context = {
        'registros': registros,
        'departamentos': departamentos,
        'departamento_seleccionado': departamento,
    }
    return render(request, 'reporte.html', context)



# Create your views here.
@login_required
@grupo_requerido('rrhh')
def validar_asistencias(request):
    sucursales = Sucursal.objects.all()
    resultados = []

    sucursal_id = request.GET.get('sucursal')
    usar_rango = request.GET.get('usar_rango') == 'on'
    fecha_str = request.GET.get('fecha')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_final_str = request.GET.get('fecha_final')

    fecha_inicio = fecha_final = None

    if sucursal_id:
        try:
            if usar_rango:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date()
            else:
                fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                fecha_final = fecha_inicio

            dias_rango = (fecha_final - fecha_inicio).days + 1
            fechas = [fecha_inicio + timedelta(days=i) for i in range(dias_rango)]

            empleados = Empleado.objects.filter(sucursal_id=sucursal_id, activo=True).select_related('sucursal')
            empleados_ids = list(empleados.values_list('id', flat=True))

            marcajes = MarcajeDepurado.objects.filter(
                empleado_id__in=empleados_ids,
                fecha__range=(fecha_inicio, fecha_final)
            )
            permisos = Permisos.objects.filter(
                empleado_id__in=empleados_ids,
                fecha_inicio__lte=fecha_final,
                fecha_final__gte=fecha_inicio
            ).select_related('tipo_permiso')

            marcajes_map = defaultdict(dict)
            for m in marcajes:
                marcajes_map[m.empleado_id][m.fecha] = m

            permisos_map = defaultdict(list)
            for p in permisos:
                permisos_map[p.empleado_id].append(p)

                ESTADO_SOLICITUD = [
                ('P', 'Pendiente'),
                ('A', 'Aprobada'),
                ('SB', 'Subsanar'),
                ('R', 'Rechazada'),
                ]
                ESTADO_MAP = dict(ESTADO_SOLICITUD)

            for empleado in empleados:
                for fecha in fechas:
                    marcaje = marcajes_map[empleado.id].get(fecha)
                    
                    permiso_contrato_vencido = next(
                        (p for p in permisos_map[empleado.id]
                        if p.fecha_inicio <= fecha <= p.fecha_final
                        and p.tipo_permiso.tipo.strip().lower() == 'contrato vencido'),
                        None
                    )

                    if permiso_contrato_vencido:
                        estado = 'JUSTIFICADO'
                        nombre_tipo = permiso_contrato_vencido.tipo_permiso.tipo
                        simbolo_permiso = permiso_contrato_vencido.tipo_permiso.simbolo
                        color = permiso_contrato_vencido.tipo_permiso.cod_color
                        estado_rh = permiso_contrato_vencido.estado_solicitud
                        estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                        entrada = simbolo_permiso
                        salida = '--:--'

                    elif marcaje:
                        estado = 'ASISTI√ì'
                        if not marcaje.entrada and marcaje.salida:
                            entrada = 'NM'
                        elif marcaje.entrada:
                            entrada = marcaje.entrada.strftime('%H:%M')
                        else:
                            entrada = '--:--'
                        salida = marcaje.salida.strftime('%H:%M') if marcaje.salida else '--:--'
                        simbolo_permiso = nombre_tipo = color = estado_rh = estado_rh_display = None

                    else:
                        if fecha.weekday() == 6:
                            estado = 'DOMINGO'
                            color = "#FFFF00"  # Color amarillo claro
                            simbolo_permiso = nombre_tipo = estado_rh = estado_rh_display = None

                            permisos_activos = [
                                p for p in permisos_map[empleado.id]
                                if p.fecha_inicio <= fecha <= p.fecha_final
                            ]
                            permiso = permisos_activos[-1] if permisos_activos else None

                            if permiso and permiso.tipo_permiso.tipo.strip().lower() == 'nuevo ingreso':
                                entrada = 'N'
                            elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'sali√≥':
                                entrada = 'S'
                            elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'contrato vencido':
                                entrada = 'CV'
                            else:
                                entrada = 'DO'
                            salida = '--:--'

                        else:
                            permisos_activos = [
                                p for p in permisos_map[empleado.id]
                                if p.fecha_inicio <= fecha <= p.fecha_final
                            ]
                            permiso = permisos_activos[-1] if permisos_activos else None

                            if permiso:
                                estado = 'JUSTIFICADO'
                                nombre_tipo = permiso.tipo_permiso.tipo
                                simbolo_permiso = permiso.tipo_permiso.simbolo
                                color = permiso.tipo_permiso.cod_color
                                estado_rh = permiso.estado_solicitud
                                estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                                entrada = simbolo_permiso
                                salida = '--:--'
                            else:
                                estado = 'FALT√ì'
                                simbolo_permiso = nombre_tipo = color = estado_rh = estado_rh_display = None
                                entrada = salida = '--:--'
                            
                    if estado == 'ASISTI√ì':
                        simbolo_excel = '‚úî'
                    elif estado == 'FALT√ì':
                        simbolo_excel = 'X'
                    elif estado == 'DOMINGO':
                        simbolo_excel = 'DO'
                    elif estado == 'JUSTIFICADO':
                        simbolo_excel = simbolo_permiso or ''
                    else:
                        simbolo_excel = ''

                    resultados.append({
                        'fecha': fecha,
                        'sucursal': empleado.sucursal.nombre,
                        'tipo_nomina': empleado.tipo_nomina,
                        'codigo': empleado.codigo,
                        'nombre': empleado.nombre,
                        'departamento': empleado.departamento,
                        'asistio': marcaje is not None,
                        'entrada': entrada,
                        'salida': salida,
                        'estado': estado,
                        'simbolo_permiso': simbolo_permiso,
                        'nombre_tipo': nombre_tipo,
                        'color': color,
                        'estado_rh': estado_rh,
                        'estado_rh_display': estado_rh_display,
                        'estado_simbolo': simbolo_excel,
                    })

        except Exception as e:
            resultados = []

    context = {
        'sucursales': sucursales,
        'resultados': resultados,
        'selected_sucursal': sucursal_id,
        'selected_fecha': fecha_str,
        'fecha_inicio': fecha_inicio_str,
        'fecha_final': fecha_final_str,
        'usar_rango': usar_rango,
    }

    if request.headers.get('HX-Request'):
        return render(request, 'partials/mostrar-asistencia.html', context)

    return render(request, 'validar_asistencia.html', context)

@login_required
@grupo_requerido('rrhh')
def crear_permiso_especial(request):
    tipo_permisos = TipoPermisos.objects.all()
    empleados = Empleado.objects.filter(activo=True)

    if request.method == 'POST':
        try:
            empleado_id = request.POST.get('empleado')
            tipo_permiso_id = request.POST.get('tipo_permiso')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_final = request.POST.get('fecha_final')
            descripcion = request.POST.get('descripcion')

            usar_hora = request.POST.get('usar_hora')
            hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            hora_final = request.POST.get('hora_final') if usar_hora else None

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            # Validaciones de existencia
            if not empleado_id or not empleado_id.isdigit():
                return render(request, 'crear_permiso_especial.html', {
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'error': "Debe seleccionar un empleado v√°lido.",
                    'datos_formulario': request.POST
                })

            if not tipo_permiso_id or not tipo_permiso_id.isdigit():
                return render(request, 'crear_permiso_especial.html', {
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'error': "Debe seleccionar un tipo de permiso v√°lido.",
                    'datos_formulario': request.POST
                })

            empleado = Empleado.objects.get(id=empleado_id)
            tipo_permiso = TipoPermisos.objects.get(id=tipo_permiso_id)

            # Validaci√≥n de fecha
            if fecha_inicio > fecha_final:
                return render(request, 'crear_permiso_especial.html', {
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final.",
                    'datos_formulario': request.POST
                })

            # Validaci√≥n de traslape
            traslape = Permisos.objects.filter(
                empleado=empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=fecha_final,
                fecha_final__gte=fecha_inicio
            ).exists()

            if traslape and not confirmar_traslape:
                return render(request, 'crear_permiso_especial.html', {
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            # Guardado
            Permisos.objects.create(
                empleado=empleado,
                tipo_permiso=tipo_permiso,
                fecha_inicio=fecha_inicio,
                fecha_final=fecha_final,
                descripcion=descripcion,
                hora_inicio=hora_inicio,
                hora_final=hora_final,
                tiene_comprobante=False,
                estado_solicitud='P',
                encargado=None
            )

            messages.success(request, "Permiso creado exitosamente.")
            return redirect('subir_comprobantes_especial')

        except Empleado.DoesNotExist:
            return render(request, 'crear_permiso_especial.html', {
                'tipo_permisos': tipo_permisos,
                'empleados': empleados,
                'error': "Empleado no v√°lido.",
            })
        except Exception as e:
            return render(request, 'crear_permiso_especial.html', {
                'tipo_permisos': tipo_permisos,
                'empleados': empleados,
                'error': f"Error al guardar: {e}"
            })

    return render(request, 'crear_permiso_especial.html', {
        'tipo_permisos': tipo_permisos,
        'empleados': empleados,
    })



@login_required
@grupo_requerido('rrhh')
def editar_permiso_especial(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)
    tipo_permisos = TipoPermisos.objects.all()
    empleados = Empleado.objects.filter(activo=True)

    if request.method == 'POST':
        try:
            empleado_id = request.POST.get('empleado')
            tipo_permiso_id = request.POST.get('tipo_permiso')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_final = request.POST.get('fecha_final')
            descripcion = request.POST.get('descripcion')
            usar_hora = request.POST.get('usar_hora')

            hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            hora_final = request.POST.get('hora_final') if usar_hora else None

            if fecha_inicio > fecha_final:
                return render(request, 'editar_permiso_especial.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final.",
                    'datos_formulario': request.POST
                })

            empleado = Empleado.objects.get(id=empleado_id)
            tipo_permiso = TipoPermisos.objects.get(id=tipo_permiso_id)

            traslape = Permisos.objects.filter(
                empleado=empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=fecha_final,
                fecha_final__gte=fecha_inicio
            ).exclude(id=permiso.id).exists()

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            if traslape and not confirmar_traslape:
                return render(request, 'editar_permiso_especial.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'empleados': empleados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            # Actualizar campos
            permiso.empleado = empleado
            permiso.tipo_permiso = tipo_permiso
            permiso.fecha_inicio = fecha_inicio
            permiso.fecha_final = fecha_final
            permiso.descripcion = descripcion
            permiso.hora_inicio = hora_inicio
            permiso.hora_final = hora_final
            permiso.encargado = None
            permiso.save()

            messages.success(request, "Permiso actualizado correctamente.")
            return redirect('subir_comprobantes_especial')

        except Empleado.DoesNotExist:
            messages.error(request, 'Empleado no v√°lido')
        except Exception as e:
            messages.error(request, f"Error al actualizar: {e}")
            return render(request, 'editar_permiso_especial.html', {
                'permiso': permiso,
                'tipo_permisos': tipo_permisos,
                'empleados': empleados,
                'error': str(e),
                'datos_formulario': request.POST
            })

    return render(request, 'editar_permiso_especial.html', {
        'permiso': permiso,
        'tipo_permisos': tipo_permisos,
        'empleados': empleados,
    })


@login_required
@grupo_requerido('rrhh')
@login_required
def eliminar_permiso_especial(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)

    if request.method == 'POST':
        try:
            permiso.delete()
            messages.success(request, "Permiso eliminado correctamente.")
            return redirect('subir_comprobantes_especial')
        except ProtectedError:
            messages.error(request, "No se puede eliminar este permiso porque tiene registros relacionados.")
            return redirect('subir_comprobantes_especial')
        except Exception as e:
            messages.error(request, f"Ocurri√≥ un error: {e}")
            return redirect('subir_comprobantes_especial')

    return render(request, 'eliminar_permiso_especial.html', {'permiso': permiso})


@login_required
@grupo_requerido('encargado')
def crear_permiso(request):
    tipo_permisos = TipoPermisos.objects.exclude(tipo__in=[
        'Especial', 'Servicios Profesionales', 'Sali√≥', 'Suspensi√≥n',
        'Incapacidad sin Seguro Social', 'Incapacidad con Seguro Social',
        'Domingo', 'Asueto', 'Nuevo Ingreso'
    ])
    encargados = Empleado.objects.filter(es_encargado=True)

    if request.method == 'POST':
        try:
            encargado_id = request.POST.get('encargado')
            empleado_id = request.POST.get('empleado')
            tipo_permiso_id = request.POST.get('tipo_permiso')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_final = request.POST.get('fecha_final')
            descripcion = request.POST.get('descripcion')

            usar_hora = request.POST.get('usar_hora')
            hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            hora_final = request.POST.get('hora_final') if usar_hora else None

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            # Validaciones de existencia
            empleado = Empleado.objects.get(id=empleado_id)
            encargado = Empleado.objects.get(id=encargado_id)
            tipo_permiso = TipoPermisos.objects.get(id=tipo_permiso_id)

            if not empleado_id or not empleado_id.isdigit():
                return render(request, 'crear_permiso.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un empleado v√°lido.",
                    'datos_formulario': request.POST
                })

            if not encargado_id or not encargado_id.isdigit():
                return render(request, 'crear_permiso.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un encargado v√°lido.",
                    'datos_formulario': request.POST
                })

            if not tipo_permiso_id or not tipo_permiso_id.isdigit():
                return render(request, 'crear_permiso.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un tipo de permiso v√°lido.",
                    'datos_formulario': request.POST
                })


            # Validaci√≥n: fecha inicio <= fecha final
            if fecha_inicio > fecha_final:
                return render(request, 'crear_permiso.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final.",
                    'datos_formulario': request.POST
                })

            # Validaci√≥n de traslape
            traslape = Permisos.objects.filter(
                empleado=empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=fecha_final,
                fecha_final__gte=fecha_inicio
            ).exists()

            if traslape and not confirmar_traslape:
                return render(request, 'crear_permiso.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            # Guardado si no hay errores
            Permisos.objects.create(
                encargado=encargado,
                empleado=empleado,
                tipo_permiso=tipo_permiso,
                fecha_inicio=fecha_inicio,
                fecha_final=fecha_final,
                descripcion=descripcion,
                hora_inicio=hora_inicio,
                hora_final=hora_final
            )

            messages.success(request, "Permiso creado exitosamente.")
            return redirect('subir_comprobantes')

        except Empleado.DoesNotExist:
            return render(request, 'crear_permiso.html', {
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': "Empleado o encargado no v√°lido.",
            })
        except Exception as e:
            return render(request, 'crear_permiso.html', {
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': f"Error al guardar: {e}"
            })

    return render(request, 'crear_permiso.html', {
        'tipo_permisos': tipo_permisos,
        'encargados': encargados,
    })


@login_required
@grupo_requerido('encargado')
def editar_permiso(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)
    tipo_permisos = TipoPermisos.objects.exclude(tipo__in=[
        'Especial', 'Servicios Profesionales', 'Sali√≥', 'Suspensi√≥n',
        'Incapacidad sin Seguro Social', 'Incapacidad con Seguro Social',
        'Domingo', 'Asueto', 'Nuevo Ingreso'
    ])
    encargados = Empleado.objects.filter(es_encargado=True)

    if request.method == 'POST':
        try:
            permiso.encargado_id = request.POST.get('encargado')
            permiso.empleado_id = request.POST.get('empleado')
            permiso.tipo_permiso_id = request.POST.get('tipo_permiso')
            permiso.fecha_inicio = request.POST.get('fecha_inicio')
            permiso.fecha_final = request.POST.get('fecha_final')
            permiso.descripcion = request.POST.get('descripcion')

            usar_hora = request.POST.get('usar_hora')
            permiso.hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            permiso.hora_final = request.POST.get('hora_final') if usar_hora else None

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            # Validar fechas
            if permiso.fecha_inicio > permiso.fecha_final:
                return render(request, 'editar_permiso.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final."
                })

            # Validar traslape excluyendo el permiso actual
            traslape = Permisos.objects.filter(
                empleado=permiso.empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=permiso.fecha_final,
                fecha_final__gte=permiso.fecha_inicio
            ).exclude(id=permiso.id).exists()

            if traslape and not confirmar_traslape:
                return render(request, 'editar_permiso.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            permiso.save()
            messages.success(request, "Permiso actualizado correctamente.")
            return redirect('subir_comprobantes')

        except Exception as e:
            return render(request, 'editar_permiso.html', {
                'permiso': permiso,
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': f"Error al actualizar: {e}"
            })

    return render(request, 'editar_permiso.html', {
        'permiso': permiso,
        'tipo_permisos': tipo_permisos,
        'encargados': encargados,
    })


@login_required
@grupo_requerido('encargado')
@login_required
def eliminar_permiso(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)

    if request.method == 'POST':
        permiso.delete()
        return redirect('subir_comprobantes')

    return render(request, 'eliminar_permiso.html', {'permiso': permiso})



@login_required
@grupo_requerido('encargado')
def crear_incapacidad(request):
    tipo_permisos = TipoPermisos.objects.filter(tipo__in=[
        'Incapacidad sin Seguro Social', 'Incapacidad con Seguro Social'
    ])
    encargados = Empleado.objects.filter(es_encargado=True, activo=True)

    if request.method == 'POST':
        try:
            encargado_id = request.POST.get('encargado')
            empleado_id = request.POST.get('empleado')
            tipo_permiso_id = request.POST.get('tipo_permiso')
            fecha_inicio = request.POST.get('fecha_inicio')
            fecha_final = request.POST.get('fecha_final')
            descripcion = request.POST.get('descripcion')

            usar_hora = request.POST.get('usar_hora')
            hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            hora_final = request.POST.get('hora_final') if usar_hora else None

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            # Validaciones de existencia
            empleado = Empleado.objects.get(id=empleado_id)
            encargado = Empleado.objects.get(id=encargado_id)
            tipo_permiso = TipoPermisos.objects.get(id=tipo_permiso_id)

            if not empleado_id or not empleado_id.isdigit():
                return render(request, 'crear_incapacidad.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un empleado v√°lido.",
                    'datos_formulario': request.POST
                })

            if not encargado_id or not encargado_id.isdigit():
                return render(request, 'crear_incapacidad.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un encargado v√°lido.",
                    'datos_formulario': request.POST
                })

            if not tipo_permiso_id or not tipo_permiso_id.isdigit():
                return render(request, 'crear_incapacidad.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "Debe seleccionar un tipo de permiso v√°lido.",
                    'datos_formulario': request.POST
                })

            # Validaci√≥n: fecha inicio <= fecha final
            if fecha_inicio > fecha_final:
                return render(request, 'crear_incapacidad.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final.",
                    'datos_formulario': request.POST
                })

            # Validaci√≥n de traslape
            traslape = Permisos.objects.filter(
                empleado=empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=fecha_final,
                fecha_final__gte=fecha_inicio
            ).exists()

            if traslape and not confirmar_traslape:
                return render(request, 'crear_incapacidad.html', {
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            # Guardado si no hay errores
            Permisos.objects.create(
                encargado=encargado,
                empleado=empleado,
                tipo_permiso=tipo_permiso,
                fecha_inicio=fecha_inicio,
                fecha_final=fecha_final,
                descripcion=descripcion,
                hora_inicio=hora_inicio,
                hora_final=hora_final
            )

            messages.success(request, "Permiso de incapacidad creado exitosamente.")
            return redirect('subir_comprobantes')

        except Empleado.DoesNotExist:
            return render(request, 'crear_incapacidad.html', {
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': "Empleado o encargado no v√°lido.",
            })
        except Exception as e:
            return render(request, 'crear_incapacidad.html', {
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': f"Error al guardar: {e}"
            })

    return render(request, 'crear_incapacidad.html', {
        'tipo_permisos': tipo_permisos,
        'encargados': encargados,
    })

@login_required
@grupo_requerido('encargado')
def editar_incapacidad(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)
    tipo_permisos = TipoPermisos.objects.filter(tipo__in=[
        'Incapacidad sin Seguro Social', 'Incapacidad con Seguro Social'
    ])
    encargados = Empleado.objects.filter(es_encargado=True)

    if request.method == 'POST':
        try:
            permiso.encargado_id = request.POST.get('encargado')
            permiso.empleado_id = request.POST.get('empleado')
            permiso.tipo_permiso_id = request.POST.get('tipo_permiso')
            permiso.fecha_inicio = request.POST.get('fecha_inicio')
            permiso.fecha_final = request.POST.get('fecha_final')
            permiso.descripcion = request.POST.get('descripcion')

            usar_hora = request.POST.get('usar_hora')
            permiso.hora_inicio = request.POST.get('hora_inicio') if usar_hora else None
            permiso.hora_final = request.POST.get('hora_final') if usar_hora else None

            confirmar_traslape = request.POST.get('confirmar_traslape') == 'true'

            # Validar fechas
            if permiso.fecha_inicio > permiso.fecha_final:
                return render(request, 'editar_incapacidad.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'error': "La fecha de inicio no puede ser posterior a la fecha final.",
                    'datos_formulario': request.POST
                })

            # Validar traslape excluyendo el permiso actual
            traslape = Permisos.objects.filter(
                empleado=permiso.empleado,
                estado_solicitud__in=['P', 'A'],
                fecha_inicio__lte=permiso.fecha_final,
                fecha_final__gte=permiso.fecha_inicio
            ).exclude(id=permiso.id).exists()

            if traslape and not confirmar_traslape:
                return render(request, 'editar_incapacidad.html', {
                    'permiso': permiso,
                    'tipo_permisos': tipo_permisos,
                    'encargados': encargados,
                    'mostrar_confirmacion': True,
                    'error': "Ya existe un permiso pendiente o aprobado que se traslapa con estas fechas.",
                    'datos_formulario': request.POST
                })

            permiso.save()
            messages.success(request, "Incapacidad actualizada correctamente.")
            return redirect('subir_comprobantes')

        except Exception as e:
            return render(request, 'editar_incapacidad.html', {
                'permiso': permiso,
                'tipo_permisos': tipo_permisos,
                'encargados': encargados,
                'error': f"Error al actualizar: {e}",
                'datos_formulario': request.POST
            })

    return render(request, 'editar_incapacidad.html', {
        'permiso': permiso,
        'tipo_permisos': tipo_permisos,
        'encargados': encargados,
    })




@login_required
@grupo_requerido('encargado')
def ficha_permiso(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)

    return render(request, 'ficha.html', {
        'solicitud': permiso,
    })

@login_required
@grupo_requerido('rrhh')
def obtener_empleados(request):
    sucursal_id = request.GET.get('sucursal_id')
    departamento = request.GET.get('departamento')

    empleados = Empleado.objects.filter(
        sucursal_id=sucursal_id,
        departamento=departamento
    ).values('id', 'nombre')

    return JsonResponse(list(empleados), safe=False)

@login_required
@grupo_requerido('rrhh')
def cargar_empleados_por_encargado(request):
    encargado_id = request.GET.get('encargado_id')
    if encargado_id:
        empleados = Empleado.objects.filter(
            encargado_asignado__encargado_id=encargado_id
        ).values('id', 'nombre')
        return JsonResponse({'empleados': list(empleados)})
    return JsonResponse({'empleados': []})

@login_required
@grupo_requerido('rrhh')
def get_empleados_por_encargado(request, encargado_id):
    encargado = get_object_or_404(Empleado, id=encargado_id, es_encargado=True)
    asignaciones = AsignacionEmpleadoEncargado.objects.select_related('empleado', 'encargado')

    asignaciones = AsignacionEmpleadoEncargado.objects.filter(encargado=encargado, empleado__activo=True).select_related('empleado')
    empleados = [a.empleado for a in asignaciones]

    return render(request, 'asignados.html', {
        'encargado': encargado,
        'empleados': empleados,
        'encargado_id': encargado.id  # üëà necesario para el template
    })

@login_required
@grupo_requerido('rrhh')
def ver_empleados_asignados(request, encargado_id):
    encargado_id = request.GET.get('encargado_id')
    if encargado_id:
        empleados = Empleado.objects.filter(
            encargado_asignado__encargado_id=encargado_id
        ).values('id', 'nombre')
        
    
@login_required
@grupo_requerido('rrhh')
def empleados_y_encargados(request):
    empleados = Empleado.objects.filter(es_encargado=False, activo=True)
    encargados = Empleado.objects.filter(es_encargado=True, activo=True)

    return render(request, 'emp_enc.html', {'empleados': empleados, 'encargados': encargados})

@login_required
@grupo_requerido('rrhh')
def convertir_a_encargado(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    empleado.es_encargado = True
    empleado.activo = True
    empleado.save()

    empleados = Empleado.objects.filter(es_encargado=False, activo=True)
    encargados = Empleado.objects.filter(es_encargado=True, activo=True)
    html = render_to_string('empleados_y_encargados.html', {
        'empleados': empleados,
        'encargados': encargados
    })
    return HttpResponse(html)

@login_required
@grupo_requerido('rrhh')
def convertir_a_empleado(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    empleados_asignados = Empleado.objects.filter(encargado_asignado__encargado=empleado, activo=True)
    
    if request.method == "POST":
        if empleados_asignados.exists() and not request.POST.get("reasignacion_completa"):
            # Mostrar formulario de reasignaci√≥n
            encargados_disponibles = Empleado.objects.filter(es_encargado=True, activo=True).exclude(id=empleado.id)
            html = render_to_string('reasignar_encargado.html', {
                'empleado': empleado,
                'empleados_asignados': empleados_asignados,
                'encargados_disponibles': encargados_disponibles
            }, request=request)
            return HttpResponse(html)
        
        # Si ya vienen los nuevos encargados en el POST, procesamos la reasignaci√≥n
        nuevo_encargado_id = request.POST.get("nuevo_encargado")
        if empleados_asignados.exists() and nuevo_encargado_id:
            nuevo_encargado = Empleado.objects.get(id=nuevo_encargado_id, es_encargado=True)
            # Actualiza todas las asignaciones
            for asignado in empleados_asignados:
                asignacion = AsignacionEmpleadoEncargado.objects.get(empleado=asignado)
                asignacion.encargado = nuevo_encargado
                asignacion.save()

        # Ahora s√≠, quita el rol
        empleado.es_encargado = False
        empleado.activo = True  # Aseg√∫rate de que el empleado siga activo
        empleado.save()

    empleados = Empleado.objects.filter(es_encargado=False, activo=True)
    encargados = Empleado.objects.filter(es_encargado=True, activo=True)
    html = render_to_string('empleados_y_encargados.html', {
        'empleados': empleados,
        'encargados': encargados
    }, request=request)
    return HttpResponse(html)


@login_required
@grupo_requerido('rrhh')
def ver_encargados(request):
    encargados = Empleado.objects.filter(es_encargado=True, activo=True)

    return render(request, 'ver_encargados.html', {'encargados': encargados})


@login_required
@grupo_requerido('rrhh')
def asignar_empleados(request, encargado_id):
    encargado = get_object_or_404(Empleado, id=encargado_id, es_encargado=True)

    departamento_seleccionado = request.POST.get('departamento', '')
    
    if request.method == 'POST' and 'empleados_ids' in request.POST:
        empleados_ids = request.POST.getlist('empleados_ids')
        for empleado_id in empleados_ids:
            AsignacionEmpleadoEncargado.objects.get_or_create(
                encargado=encargado,
                empleado_id=empleado_id
            )

    # ‚úÖ Incluye a encargados, excepto el encargado actual
    empleados_disponibles = Empleado.objects.filter(
        encargado_asignado__isnull=True, activo=True
    )

    if departamento_seleccionado:
        empleados_disponibles = empleados_disponibles.filter(departamento=departamento_seleccionado)

    departamentos_qs = Empleado.objects.filter(
        encargado_asignado__isnull=True
    ).values_list('departamento', flat=True)

    departamentos_limpios = {}
    for depto in departamentos_qs:
        key = depto.strip().upper()
        if key not in departamentos_limpios:
            departamentos_limpios[key] = depto.strip()

    departamentos = sorted(departamentos_limpios.values())

    html = render_to_string('asignar_empleados.html', {
        'encargado': encargado,
        'empleados': empleados_disponibles,
        'departamentos': departamentos,
        'departamento_seleccionado': departamento_seleccionado,
    }, request=request)

    return HttpResponse(html)

@login_required
@grupo_requerido('rrhh')
def quitar_empleado_asignado(request, encargado_id, empleado_id):
    encargado = get_object_or_404(Empleado, id=encargado_id, es_encargado=True)
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == "POST":
        asignacion = AsignacionEmpleadoEncargado.objects.filter(encargado=encargado, empleado=empleado)
        if asignacion.exists():
            asignacion.delete()
        # Recarga la lista actualizada y retorna solo el fragmento HTML
        asignaciones = AsignacionEmpleadoEncargado.objects.filter(encargado=encargado).select_related('empleado')
        empleados = [a.empleado for a in asignaciones]
        return render(request, 'asignados.html', {
            'encargado': encargado,
            'empleados': empleados,
        })
    return HttpResponse(status=405)  # M√©todo no permitido si no es POST

@login_required
@grupo_requerido('rrhh')
def solicitud_rh(request):
    estado = 'P'
    permisos = Permisos.objects.filter(Q(estado_solicitud=estado) | Q(estado_solicitud='SB')).order_by('-fecha_solicitud')

    context = []
    for permiso in permisos:
        comprobante = PermisoComprobante.objects.filter(permiso=permiso).first()
        context.append({
            'permiso': permiso,
            'estado_rh_display': permiso.get_estado_solicitud_display(),  # ‚úÖ m√°s limpio
            'comprobante': comprobante.comprobante.url if comprobante else None,
        })

    return render(request, 'solicitudes_rh.html', {'permisos': context},)


@login_required
@grupo_requerido('encargado')
def vista_solicitudes_encargado(request):
    try:
        encargado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        return render(request, 'sin_permisos.html')
    permisos = Permisos.objects.filter(encargado=encargado.id, estado_solicitud__in=['P', 'SB']).order_by('-fecha_solicitud')

    context = []
    for permiso in permisos:
        comprobante = PermisoComprobante.objects.filter(permiso=permiso).first()
        context.append({
            'permiso': permiso,
            'estado_rh_display': permiso.get_estado_solicitud_display(),
            'comprobante': comprobante.comprobante.url if comprobante else None,
        })
    return render(request, 'solicitudes_encargado.html', {'solicitudes': context})

@login_required
@grupo_requerido('encargado')
def ver_historial_encargado(request):
    encargado = Empleado.objects.get(user=request.user)
    permisos = Permisos.objects.filter(encargado=encargado.id, estado_solicitud__in=['A', 'R']).order_by('-fecha_solicitud')
    context=[]
    for permiso in permisos:
        comprobante = PermisoComprobante.objects.filter(permiso=permiso).first()
        historial = GestionPermisoDetalle.objects.filter(solicitud=permiso).order_by('-fecha').first()
        context.append({
            'permiso': permiso,
            'comprobante': comprobante.comprobante.url if comprobante else None,
            'estado_rh_display': historial.solicitud.get_estado_solicitud_display(),
            'historial': historial,
        })

    return render(request, 'historial_solicitudes_encargado.html', {'solicitudes': context})


@login_required
@grupo_requerido('rrhh')
def subir_comprobante_especial(request):
    solicitudes_raw = Permisos.objects.filter(
        Q(tiene_comprobante=False) | Q(estado_solicitud='SB', pendiente_subsanar=True)
    ).exclude(encargado__isnull=False)  # Excluye todos los que ya tienen encargado

    solicitudes_raw = solicitudes_raw.order_by('-fecha_solicitud')

    solicitudes = []
    for permiso in solicitudes_raw:
        solicitudes.append({
            'permiso': permiso,
            'estado_rh_display': permiso.get_estado_solicitud_display(),
        })

    return render(request, 'subir_comprobantes_especial.html', {
        'solicitudes': solicitudes
    })



@login_required
@grupo_requerido('encargado')
def subir_comprobante(request):
    try:
        encargado = Empleado.objects.get(user=request.user)
    except Empleado.DoesNotExist:
        return render(request, 'sin_permisos.html')

    permisos = Permisos.objects.filter(
        Q(encargado=encargado, tiene_comprobante=False) |
        Q(encargado=encargado, estado_solicitud='SB', pendiente_subsanar=True)
    ).order_by('-fecha_solicitud')

    solicitudes = []
    for permiso in permisos:
        historial = GestionPermisoDetalle.objects.filter(solicitud=permiso).order_by('-fecha').first()
        solicitudes.append({
            'permiso': permiso,
            'estado_rh_display': permiso.get_estado_solicitud_display(),
            'historial': historial,
        })

    return render(request, 'subir_comprobantes.html', {
        'solicitudes': solicitudes,
    })

@login_required
@grupo_requerido('encargado')
def formulario_comprobantes(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)
    comprobante_qs = PermisoComprobante.objects.filter(permiso=permiso)
    if request.method == 'POST':
        form = SubirComprobanteForm(request.POST, request.FILES, instance=comprobante_qs.first() if comprobante_qs.exists() else None)

        if form.is_valid():
            comprobante = form.save(commit=False)
            comprobante.permiso = permiso
            comprobante.save()
            permiso.tiene_comprobante = True
            permiso.pendiente_subsanar = False
            permiso.save()
            html = render_to_string('act_fila_comp.html', {"permiso": permiso})
            return HttpResponse(html)
    else:
        form = SubirComprobanteForm(instance=comprobante_qs.first() if comprobante_qs.exists() else None)
    
    return render(request, "formulario.html", {"form": form, "permiso": permiso})

@login_required
@grupo_requerido('rrhh')
def crear_usuario(request):
    encargados = Empleado.objects.filter(es_encargado=True)

    if request.method == 'POST':
        encargado_id = request.POST.get('encargado')
        password = request.POST.get('password')
        email = request.POST.get('email')

        try:
            empleado = Empleado.objects.get(id=encargado_id)
        except Empleado.DoesNotExist:
            messages.error(request, "Encargado no v√°lido.")
            return redirect('crear_usuario')

        if empleado.user:
            messages.error(request, f"El encargado {empleado.nombre} ya tiene un usuario asignado.")
            return redirect('crear_usuario')

        nombres = empleado.nombre.strip().lower().split()

        if len(nombres) >= 2:
            username = f"{nombres[0]}{nombres[-2]}"
            first_name = nombres[0].capitalize()
            last_name = nombres[-2].capitalize()
        else:
            username = nombres[0]
            first_name = nombres[0].capitalize()
            last_name = ""

        user = User.objects.create_user(username=username, email=email, password=password)
        user.first_name = first_name
        user.last_name = last_name
        grupo_encargado, created = Group.objects.get_or_create(name="encargado")
        user.groups.add(grupo_encargado)
        user.save()

        empleado.user = user
        empleado.save()

        messages.success(request, f"Usuario creado exitosamente para {empleado.nombre}.")
        return redirect('crear_usuario')

    return render(request, 'crear_usuario.html', {'encargados': encargados})

@login_required
@grupo_requerido('rrhh')
def ver_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'listar_usuarios.html', {'usuarios': usuarios})

@login_required
@grupo_requerido('rrhh')
def modal_solicitud(request, permiso_comprobante_id):
    permiso_comprobante = get_object_or_404(PermisoComprobante, id=permiso_comprobante_id)
    # permiso = get_object_or_404(Permisos, permisos_id=permiso_id )
    return render(request, 'partials/modal.html', {'permiso_comprobante': permiso_comprobante,})

@login_required
@grupo_requerido('rrhh')
def accion_solicitud(request):
    
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        accion = request.POST.get('accion_realizada')  # 'aprobar' o 'rechazar'
        revisado_por = request.POST.get('revisada_por')
        comentario = request.POST.get('comentarios', '')
        aprobacion_gerencial = request.POST.get('aprobacion_gerencial')

        # 1. Busca la solicitud
        solicitud = get_object_or_404(Permisos, id=solicitud_id)
        
        # 2. Define el nuevo estado
        if accion == 'A':
            nuevo_estado = 'A'
            if aprobacion_gerencial:
                solicitud.aprobacion_gerencial = True
        elif accion == 'R':
            nuevo_estado = 'R'
        elif accion == 'SB':
            nuevo_estado = 'SB'
            solicitud.pendiente_subsanar = True
        else:
            return HttpResponse("Acci√≥n no v√°lida", status=400)

        # 3. Guarda el detalle
        GestionPermisoDetalle.objects.create(
            solicitud=solicitud,
            accion_realizada=nuevo_estado,
            revisada_por=revisado_por,
            comentarios=comentario
        )

        # 4. Actualiza el estado de la solicitud
        solicitud.estado_solicitud = nuevo_estado
        solicitud.save()



        # 5. (Opcional) Retorna un mensaje, puedes actualizar la tabla con HTMX
        return redirect('solicitud_rh')

    return HttpResponse("M√©todo no permitido", status=405)

@login_required
@grupo_requerido('rrhh')
def ver_historial_solicitudes(request):
    historial = GestionPermisoDetalle.objects.filter(solicitud__estado_solicitud__in=['A', 'R', 'SB']).order_by('-fecha')
    context = []
    for h in historial:
        comprobante_obj = PermisoComprobante.objects.filter(permiso=h.solicitud).first()
        comprobante_url = comprobante_obj.comprobante.url if comprobante_obj else None

        context.append({
            'detalle': h,
            'estado_rh_display': h.solicitud.get_estado_solicitud_display(),
            'comprobante_url': comprobante_url,
        })
    return render(request, 'historial_solicitudes.html', {'solicitudes': context})

def eliminar_permiso_H(request, permiso_id):
    permiso = get_object_or_404(Permisos, id=permiso_id)

    if request.method == 'POST':
        permiso.delete()
        return redirect('historial_solicitudes')  # o donde quieras redirigir

    return render(request, 'eliminar_permiso.html', {'permiso': permiso})


def cargar_login(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'home'

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            # üëá Verificamos si el usuario debe cambiar su contrase√±a
            if hasattr(user, 'userprofile') and user.userprofile.must_change_password:
                return redirect('cambiar_password')  # o la URL que definas

            return redirect(next_url)
        else:
            return render(request, 'login.html', {
                'error': 'Usuario o contrase√±a incorrectos',
                'next': next_url
            })

    return render(request, 'login.html', {'next': next_url})

class ForzarCambioPasswordView(PasswordChangeView):
    template_name = 'cambiar_password.html'
    success_url = reverse_lazy('home')

    def dispatch(self, request, *args, **kwargs):
        # Opcional: solo permite acceso si debe cambiar password
        if not request.user.is_authenticated or not request.user.userprofile.must_change_password:
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Cambia el flag
        self.request.user.userprofile.must_change_password = False
        self.request.user.userprofile.save()

        # A√±ade un mensaje de √©xito
        messages.success(self.request, '¬°Contrase√±a cambiada correctamente!')
        return response




def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
@grupo_requerido('encargado')
def ver_a_cargo(request):
    return render(request, 'ver.html')

@login_required
@grupo_requerido('rrhh')
def ausencias_encargado(request):
    encargado = Empleado.objects.filter(es_encargado=True)
    enc_id = request.GET.get('encargado')
    fecha_str = request.GET.get('fecha')

    
    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha = date.today()
    else:
        fecha = date.today()

    empleados = []
    encargado_seleccionado = None

    if enc_id:
        encargado_seleccionado = get_object_or_404(Empleado, id=enc_id)
        asignaciones = AsignacionEmpleadoEncargado.objects.filter(encargado=enc_id).order_by('empleado__nombre')
        empleados_asignados = [a.empleado for a in asignaciones]

        for empleado in empleados_asignados:
            tiene_marcaje = MarcajeDepurado.objects.filter(empleado=empleado, fecha=fecha).exists()
            permiso_justificado = Permisos.objects.filter(
                empleado=empleado, 
                estado_solicitud__in=['A', 'SB'],
                fecha_inicio__lte=fecha,
                fecha_final__gte=fecha
            ).select_related('tipo_permiso').first()

            if tiene_marcaje:
                estado = 'ASISTI√ì'
                simbolo_permiso = None
                color = None
                estado_rh = None
            elif fecha.weekday() == 6:  # 6 = Domingo
                estado = 'DOMINGO'
                simbolo_permiso = None
                color = "#00f7ff"
                estado_rh = None                
            elif permiso_justificado:
                estado = 'JUSTIFICADO'
                simbolo_permiso = permiso_justificado.tipo_permiso.simbolo
                color = permiso_justificado.tipo_permiso.cod_color
                estado_rh = permiso_justificado.estado_solicitud                
            else:
                estado = 'FALT√ì'
                simbolo_permiso = None
                color = None
                estado_rh = None
            
            empleados.append({
                'codigo': empleado.codigo,
                'nombre': empleado.nombre,
                'departamento': empleado.departamento,
                'sucursal': empleado.sucursal.nombre,
                'estado': estado,
                'simbolo_permiso': simbolo_permiso,
                'color': color,
                'estado_rh': estado_rh,
                
            })

    return render(request, 'ausencias_encargado.html', {
        'fecha': fecha,
        'encargado': encargado,
        'empleados': empleados,
        'encargado_seleccionado': encargado_seleccionado,
        'enc_id': enc_id,
    })

@login_required
@grupo_requerido('encargado')
def asistencias_encargado(request):
    user = request.user
    try:
        encargado = Empleado.objects.get(user=user, es_encargado=True)
    except Empleado.DoesNotExist:
        return render(request, 'sin_permisos.html')

    usar_rango = request.GET.get('usar_rango') == 'on'

    fecha_str = request.GET.get('fecha')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_final_str = request.GET.get('fecha_final')

    if usar_rango:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            fecha_inicio = date.today()
        try:
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            fecha_final = fecha_inicio
    else:
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            fecha_inicio = date.today()
        fecha_final = fecha_inicio

    dias_rango = (fecha_final - fecha_inicio).days + 1
    fechas = [fecha_inicio + timedelta(days=i) for i in range(dias_rango)]

    ESTADO_SOLICITUD = [
    ('P', 'Pendiente'),
    ('A', 'Aprobada'),
    ('SB', 'Subsanar'),
    ('R', 'Rechazada'),
    ]
    ESTADO_MAP = dict(ESTADO_SOLICITUD)

    asignaciones = AsignacionEmpleadoEncargado.objects.filter(
        encargado=encargado,
        empleado__activo=True
    ).select_related('empleado', 'empleado__sucursal')

    empleados_asignados = [a.empleado for a in asignaciones]
    empleados_ids = [e.id for e in empleados_asignados]

    marcajes = MarcajeDepurado.objects.filter(
        empleado_id__in=empleados_ids,
        fecha__range=(fecha_inicio, fecha_final)
    )
    permisos = Permisos.objects.filter(
        empleado_id__in=empleados_ids,
        fecha_inicio__lte=fecha_final,
        fecha_final__gte=fecha_inicio
    ).select_related('tipo_permiso')

    marcajes_map = defaultdict(dict)
    for m in marcajes:
        marcajes_map[m.empleado_id][m.fecha] = m

    permisos_map = defaultdict(list)
    for p in permisos:
        permisos_map[p.empleado_id].append(p)

    empleados = []

    for empleado in empleados_asignados:
        for fecha in fechas:
            marcaje = marcajes_map[empleado.id].get(fecha)

            # üîπ Primero: verificar si tiene contrato vencido
            permiso_contrato_vencido = next(
                (p for p in permisos_map[empleado.id]
                if p.fecha_inicio <= fecha <= p.fecha_final
                and p.tipo_permiso.tipo.strip().lower() == 'contrato vencido'),
                None
            )

            if permiso_contrato_vencido:
                estado = 'JUSTIFICADO'
                nombre_tipo = permiso_contrato_vencido.tipo_permiso.tipo
                simbolo_permiso = permiso_contrato_vencido.tipo_permiso.simbolo
                color = permiso_contrato_vencido.tipo_permiso.cod_color or '#fbbf24'
                estado_rh = permiso_contrato_vencido.estado_solicitud
                estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                entrada = simbolo_permiso or '--'
                salida = '--:--'

            elif marcaje:
                estado = 'ASISTI√ì'
                entrada = marcaje.entrada.strftime('%H:%M') if marcaje.entrada else '--:--'
                salida = marcaje.salida.strftime('%H:%M') if marcaje.salida else '--:--'
                simbolo_permiso = nombre_tipo = estado_rh = estado_rh_display = None
                color = '#38c172'  # Verde

            else:
                if fecha.weekday() == 6:
                    estado = 'DOMINGO'
                    permisos_activos = [
                        p for p in permisos_map[empleado.id]
                        if p.fecha_inicio <= fecha <= p.fecha_final
                    ]
                    permiso = permisos_activos[-1] if permisos_activos else None

                    if permiso:
                        nombre_tipo = permiso.tipo_permiso.tipo
                        simbolo_permiso = permiso.tipo_permiso.simbolo
                        color = permiso.tipo_permiso.cod_color or "#FFFF00"
                        estado_rh = permiso.estado_solicitud
                        estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)

                        if nombre_tipo.strip().lower() == 'nuevo ingreso':
                            entrada = 'N'
                        elif nombre_tipo.strip().lower() == 'sali√≥':
                            entrada = 'S'
                        elif nombre_tipo.strip().lower() == 'contrato vencido':
                            entrada = 'CV'
                        else:
                            entrada = 'DO'
                    else:
                        # Domingo sin permiso
                        nombre_tipo = simbolo_permiso = estado_rh = estado_rh_display = None
                        color = "#FFFF00"
                        entrada = 'DO'
                        salida = '--:--'

                else:
                    permisos_validos = [
                        p for p in permisos_map[empleado.id]
                        if p.fecha_inicio <= fecha <= p.fecha_final
                    ]
                    permiso = permisos_validos[-1] if permisos_validos else None

                    if permiso:
                        estado = 'JUSTIFICADO'
                        nombre_tipo = permiso.tipo_permiso.tipo
                        simbolo_permiso = permiso.tipo_permiso.simbolo
                        color = permiso.tipo_permiso.cod_color or '#fbbf24'  # Amarillo por defecto
                        estado_rh = permiso.estado_solicitud
                        estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                        entrada = simbolo_permiso or '--'
                        salida = '--:--'
                    else:
                        estado = 'FALT√ì'
                        entrada = salida = '--:--'
                        simbolo_permiso = nombre_tipo = estado_rh = estado_rh_display = None
                        color = '#e3342f'  # Rojo

            empleados.append({
                'fecha': fecha,
                'codigo': empleado.codigo,
                'nombre': empleado.nombre,
                'departamento': empleado.departamento,
                'sucursal': empleado.sucursal.nombre,
                'tipo_nomina': empleado.tipo_nomina,
                'estado': estado,
                'simbolo_permiso': simbolo_permiso,
                'nombre_tipo': nombre_tipo,
                'color': color,
                'estado_rh': estado_rh,
                'estado_rh_display': estado_rh_display,
                'entrada': entrada,
                'salida': salida,
            })

    empleados.sort(key=lambda x: (x['codigo'], x['fecha']))

    return render(request, 'asistencias_encargado.html', {
        'fecha': fecha_inicio,
        'fecha_inicio': fecha_inicio,
        'fecha_final': fecha_final,
        'usar_rango': usar_rango,
        'encargado': encargado,
        'empleados': empleados,
    })


@login_required
@grupo_requerido('encargado')
def exportar_asistencias_encargado_excel(request):
    user = request.user
    try:
        encargado = Empleado.objects.get(user=user, es_encargado=True)
    except Empleado.DoesNotExist:
        return HttpResponse("No tiene permisos para exportar", status=403)

    usar_rango = request.GET.get('usar_rango') == 'on'
    fecha_str = request.GET.get('fecha')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_final_str = request.GET.get('fecha_final')

    if usar_rango:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        except:
            return HttpResponse("Rango de fechas inv√°lido", status=400)
        try:
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date()
        except:
            fecha_final = fecha_inicio
    else:
        try:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except:
            return HttpResponse("Fecha inv√°lida", status=400)
        fecha_final = fecha_inicio

    dias_rango = (fecha_final - fecha_inicio).days + 1

    asignaciones = AsignacionEmpleadoEncargado.objects.filter(
        encargado=encargado, empleado__activo=True
    ).select_related('empleado')
    empleados = [a.empleado for a in asignaciones]

    ESTADO_SOLICITUD = [
        ('P', 'Pendiente'),
        ('A', 'Aprobada'),
        ('SB', 'Subsanar'),
        ('R', 'Rechazada'),
    ]
    ESTADO_MAP = dict(ESTADO_SOLICITUD)
    
    resultados = []

    for empleado in empleados:
        for i in range(dias_rango):
            fecha = fecha_inicio + timedelta(days=i)

            # Obtener marcaje
            marcaje = MarcajeDepurado.objects.filter(empleado=empleado, fecha=fecha).first()
            # Obtener permisos que cubren la fecha
            permisos_validos = list(
                Permisos.objects.filter(
                    empleado=empleado,
                    fecha_inicio__lte=fecha,
                    fecha_final__gte=fecha
                )
                .select_related('tipo_permiso')
                .order_by('fecha_inicio', 'id')
            )
            permiso = permisos_validos[-1] if permisos_validos else None

            # 1Ô∏è‚É£ Contrato vencido primero
            permiso_contrato_vencido = next(
                (p for p in permisos_validos if p.tipo_permiso.tipo.strip().lower() == 'contrato vencido'),
                None
            )
            if permiso_contrato_vencido:
                estado = 'JUSTIFICADO'
                entrada = permiso_contrato_vencido.tipo_permiso.simbolo or '--'
                salida = '--:--'
                nombre_tipo = permiso_contrato_vencido.tipo_permiso.tipo
                color = permiso_contrato_vencido.tipo_permiso.cod_color.lstrip('#') if permiso_contrato_vencido.tipo_permiso.cod_color else 'fbbf24'
                estado_rh_display = ESTADO_MAP.get(permiso_contrato_vencido.estado_solicitud, permiso_contrato_vencido.estado_solicitud)
                simbolo = permiso_contrato_vencido.tipo_permiso.simbolo

            # 2Ô∏è‚É£ Asisti√≥
            elif marcaje:
                estado = 'ASISTI√ì'
                entrada = marcaje.entrada.strftime('%H:%M') if marcaje.entrada else '--:--'
                salida = marcaje.salida.strftime('%H:%M') if marcaje.salida else '--:--'
                nombre_tipo = estado_rh_display = simbolo = ''
                color = '38c172'  # Verde

            # 3Ô∏è‚É£ Domingo con prioridad sobre justificado
            elif fecha.weekday() == 6:
                estado = 'DOMINGO'
                color = "00f7ff"

                # Aqu√≠ usamos permisos_validos, que ya es la lista de permisos del empleado
                permisos_activos = [
                    p for p in permisos_validos
                    if p.fecha_inicio <= fecha <= p.fecha_final
                ]
                permiso = permisos_activos[-1] if permisos_activos else None

                if permiso and permiso.tipo_permiso.tipo.strip().lower() == 'nuevo ingreso':
                    entrada = 'N'
                elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'sali√≥':
                    entrada = 'S'
                elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'contrato vencido':
                    entrada = 'CV'
                else:
                    entrada = 'DO'

                salida = '--:--'
                simbolo_permiso = estado_rh = estado_rh_display = None
                nombre_tipo = permiso.tipo_permiso.tipo if permiso else None


            # 4Ô∏è‚É£ Justificado normal
            elif permiso:
                estado = 'JUSTIFICADO'
                entrada = permiso.tipo_permiso.simbolo or '--'
                salida = '--:--'
                nombre_tipo = permiso.tipo_permiso.tipo
                color = permiso.tipo_permiso.cod_color.lstrip('#') if permiso.tipo_permiso.cod_color else 'fbbf24'
                estado_rh_display = ESTADO_MAP.get(permiso.estado_solicitud, permiso.estado_solicitud)
                simbolo = permiso.tipo_permiso.simbolo

            # 5Ô∏è‚É£ Falt√≥
            else:
                estado = 'FALT√ì'
                entrada = salida = '--:--'
                nombre_tipo = ''
                estado_rh_display = ''
                color = 'e3342f'
                simbolo = ''

            # S√≠mbolo de asistencia
            if estado == 'ASISTI√ì':
                estado_simbolo = '‚úî'
            elif estado == 'FALT√ì':
                estado_simbolo = 'X'
            elif estado == 'DOMINGO':
                if entrada == 'N':
                    estado_simbolo = 'Nuevo Ingreso'
                elif entrada == 'S':
                    estado_simbolo = 'Sali√≥'
                elif entrada == 'CV':
                    estado_simbolo = 'Contrato Vencido'
                else:
                    estado_simbolo = 'Domingo'
            elif estado == 'JUSTIFICADO':
                estado_simbolo = nombre_tipo or ''
            else:
                estado_simbolo = ''

            resultados.append({
                'fecha': fecha.strftime('%d/%m/%Y'),
                'sucursal': empleado.sucursal.nombre,
                'tipo_nomina': empleado.tipo_nomina,
                'codigo': empleado.codigo,
                'nombre': empleado.nombre,
                'departamento': empleado.departamento,
                'entrada': entrada,
                'salida': salida,
                'estado_rh_display': estado_rh_display,
                'estado_simbolo': estado_simbolo,
                'color': color,
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = [
        "Fecha", "Sucursal", "N√≥mina", "C√≥digo", "Nombre", "Departamento",
        "Marca Entrada", "Marca Salida", "Estado RH", "Asistencia"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in resultados:
        ws.append([
            row['fecha'], row['sucursal'], row['tipo_nomina'], row['codigo'],
            row['nombre'], row['departamento'], row['entrada'], row['salida'],
            row['estado_rh_display'], row['estado_simbolo']
        ])
        ws.cell(row=ws.max_row, column=10).fill = PatternFill(start_color=row['color'], end_color=row['color'], fill_type='solid')

    filename = f"asistencia_{encargado.nombre}_{fecha_inicio.strftime('%d-%m-%Y')}"
    if usar_rango and fecha_inicio != fecha_final:
        filename += f"_a_{fecha_final.strftime('%d-%m-%Y')}"
    filename += ".xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
@grupo_requerido('rrhh')
def enviar_ausencias(request):
    if request.method == "POST":
        fecha_str = request.POST.get('fecha')
        encargado_id = request.POST.get('encargado_id')
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        encargado = get_object_or_404(Empleado, id=encargado_id)

        # Obtiene empleados asignados
        asignaciones = AsignacionEmpleadoEncargado.objects.filter(encargado=encargado)
        empleados = [a.empleado for a in asignaciones]
        ausentes = []
        for empleado in empleados:
            if not MarcajeDepurado.objects.filter(empleado=empleado, fecha=fecha).exists():
                ausentes.append(empleado)
        
        # Armado del cuerpo del correo
        lista_ausentes = "\n".join([f"{e.codigo} - {e.nombre} ({e.departamento})" for e in ausentes]) or "No hubo ausentes."
        subject = f"Reporte de Ausencias {fecha.strftime('%d/%m/%Y')}"
        mensaje = f"""Hola {encargado.nombre},

Estos son los empleados con ausencia el {fecha.strftime('%d/%m/%Y')}:

{lista_ausentes}

Este es un mensaje autom√°tico.
"""
        destinatario = encargado.user.email if encargado.user and encargado.user.email else None
        if destinatario:
            send_mail(
                subject,
                mensaje,
                'hello@demomailtrap.co',  # Cambia esto por tu correo configurado
                [destinatario],
                fail_silently=False,
            )
        # Puedes agregar mensajes de √©xito/fallo con Django messages si gustas
        return redirect('ausencias_encargado')  # O a donde quieras regresar

    # Si entran por GET
    return redirect('ausencias_encargado')

def error_404(request, exception):
    return render(request, '404.html', status=404)

def error_500(request):
    return render(request, '500.html', status=500)

@login_required
@grupo_requerido('rrhh')
def fecha_corte(request):
    if request.method == 'POST':
        anio = request.POST.get('anio')
        mes = request.POST.get('mes')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_final = request.POST.get('fecha_final')
        fecha_corte = request.POST.get('fecha_corte')

        GestionFechaCorte.objects.create(
            anio=anio,
            mes=mes,
            fecha_inicio=fecha_inicio,
            fecha_final=fecha_final,
            fecha_corte=fecha_corte,
        )

        return redirect('listar_fecha_corte')
    return render(request, 'gestion_fecha_corte.html')

@login_required
@grupo_requerido('rrhh')
def listar_fechas_corte(request):
    fecha_cortes = GestionFechaCorte.objects.all()
    return render(request, 'listar_fecha_corte.html', {'fecha_cortes': fecha_cortes})

def sin_permiso(request):
    return render(request, 'sin_permiso.html')

@login_required
@grupo_requerido('rrhh')
def reporte_asistencia(request):
    fecha_str = request.GET.get('fecha')
    departamento_seleccionado = request.GET.get('departamento', '')

    if fecha_str:
        try:
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha = date.today()
    else:
        fecha = date.today()

    # Obtener todos los departamentos posibles ese d√≠a
    departamentos_qs = MarcajeDepurado.objects.filter(fecha=fecha).select_related('empleado')\
        .values_list('empleado__departamento', flat=True).distinct()

    departamentos_limpios = {}
    for depto in departamentos_qs:
        if depto:
            key = depto.strip().upper()
            if key not in departamentos_limpios:
                departamentos_limpios[key] = depto.strip()
    departamentos = sorted(departamentos_limpios.values())

    # Consulta principal con select_related
    asistencia = MarcajeDepurado.objects.filter(fecha=fecha).select_related('empleado')
    if departamento_seleccionado:
        asistencia = asistencia.filter(empleado__departamento=departamento_seleccionado)

    return render(request, 'reporte_asistencia.html', {
        'asistencia': asistencia.order_by('entrada'),
        'fecha': fecha,
        'departamentos': departamentos,
        'departamento_seleccionado': departamento_seleccionado,
    })



@login_required
@grupo_requerido('rrhh')
def exportar_asistencias_excel(request):
    sucursal_id = request.GET.get('sucursal')
    usar_rango = request.GET.get('usar_rango') == 'on'
    fecha_str = request.GET.get('fecha')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_final_str = request.GET.get('fecha_final')

    if not sucursal_id:
        return HttpResponse("Par√°metro 'sucursal' requerido", status=400)

    try:
        if usar_rango:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date()
        else:
            fecha_inicio = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            fecha_final = fecha_inicio
    except (ValueError, TypeError):
        return HttpResponse("Fechas inv√°lidas", status=400)

    dias_rango = (fecha_final - fecha_inicio).days + 1
    fechas = [fecha_inicio + timedelta(days=i) for i in range(dias_rango)]

    empleados = Empleado.objects.filter(sucursal_id=sucursal_id, activo=True).select_related('sucursal')
    empleados_ids = list(empleados.values_list('id', flat=True))

    marcajes = MarcajeDepurado.objects.filter(
        empleado_id__in=empleados_ids,
        fecha__range=(fecha_inicio, fecha_final)
    )
    permisos = Permisos.objects.filter(
        empleado_id__in=empleados_ids,
        fecha_inicio__lte=fecha_final,
        fecha_final__gte=fecha_inicio
    ).select_related('tipo_permiso')

    marcajes_map = defaultdict(dict)
    for m in marcajes:
        marcajes_map[m.empleado_id][m.fecha] = m

    permisos_map = defaultdict(list)
    for p in permisos:
        permisos_map[p.empleado_id].append(p)

    ESTADO_SOLICITUD = [
        ('P', 'Pendiente'),
        ('A', 'Aprobada'),
        ('SB', 'Subsanar'),
        ('R', 'Rechazada'),
    ]
    ESTADO_MAP = dict(ESTADO_SOLICITUD)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    headers = [
        "Fecha", "Sucursal", "N√≥mina", "C√≥digo", "Nombre", "Departamento",
        "Marca Entrada", "Marca Salida", "Estado RH", "Asistencia"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for empleado in empleados:
        for fecha in fechas:
            marcaje = marcajes_map[empleado.id].get(fecha)

            # üîπ Primero: verificar si tiene contrato vencido
            permiso_contrato_vencido = next(
                (p for p in permisos_map[empleado.id]
                if p.fecha_inicio <= fecha <= p.fecha_final
                and p.tipo_permiso.tipo.strip().lower() == 'contrato vencido'),
                None
            )

            if permiso_contrato_vencido:
                estado = 'JUSTIFICADO'
                nombre_tipo = permiso_contrato_vencido.tipo_permiso.tipo
                simbolo_permiso = permiso_contrato_vencido.tipo_permiso.simbolo
                color = permiso_contrato_vencido.tipo_permiso.cod_color.lstrip('#') if permiso_contrato_vencido.tipo_permiso.cod_color else 'fbbf24'
                estado_rh = permiso_contrato_vencido.estado_solicitud
                estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                entrada = simbolo_permiso or '--'
                salida = '--:--'

            elif marcaje:
                estado = 'ASISTI√ì'
                if not marcaje.entrada and marcaje.salida:
                    entrada = 'NM'
                elif marcaje.entrada:
                    entrada = marcaje.entrada.strftime('%H:%M')
                else:
                    entrada = '--:--'
                salida = marcaje.salida.strftime('%H:%M') if marcaje.salida else '--:--'
                simbolo_permiso = nombre_tipo = estado_rh = estado_rh_display = None
                color = '38c172'  # Verde

            elif fecha.weekday() == 6:
                estado = 'DOMINGO'
                color = "00f7ff"
                permisos_activos = [
                    p for p in permisos_map[empleado.id]
                    if p.fecha_inicio <= fecha <= p.fecha_final
                ]
                permiso = permisos_activos[-1] if permisos_activos else None

                if permiso and permiso.tipo_permiso.tipo.strip().lower() == 'nuevo ingreso':
                    entrada = 'N'
                elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'sali√≥':
                    entrada = 'S'
                elif permiso and permiso.tipo_permiso.tipo.strip().lower() == 'contrato vencido':
                    entrada = 'CV'
                else:
                    entrada = 'DO'
                salida = '--:--'
                simbolo_permiso = estado_rh = estado_rh_display = None
                nombre_tipo = permiso.tipo_permiso.tipo if permiso else None

            else:
                permisos_validos = [
                    p for p in permisos_map[empleado.id]
                    if p.fecha_inicio <= fecha <= p.fecha_final
                ]
                permiso = permisos_validos[-1] if permisos_validos else None

                if permiso:
                    estado = 'JUSTIFICADO'
                    nombre_tipo = permiso.tipo_permiso.tipo
                    simbolo_permiso = permiso.tipo_permiso.simbolo
                    color = permiso.tipo_permiso.cod_color.lstrip('#') if permiso.tipo_permiso.cod_color else 'fbbf24'
                    estado_rh = permiso.estado_solicitud
                    estado_rh_display = ESTADO_MAP.get(estado_rh, estado_rh)
                    entrada = simbolo_permiso or '--'
                    salida = '--:--'
                else:
                    estado = 'FALT√ì'
                    entrada = salida = '--:--'
                    color = 'e3342f'
                    simbolo_permiso = nombre_tipo = estado_rh = estado_rh_display = None

            # S√≠mbolo para Excel
            if estado == 'ASISTI√ì':
                estado_simbolo = '‚úî'
            elif estado == 'FALT√ì':
                estado_simbolo = 'X'
            elif estado == 'DOMINGO':
                if nombre_tipo == 'Nuevo Ingreso':
                    estado_simbolo = 'Nuevo Ingreso'
                elif nombre_tipo == 'Sali√≥':
                    estado_simbolo = 'Sali√≥'
                elif nombre_tipo == 'Contrato Vencido':
                    estado_simbolo = 'Contrato Vencido'
                else:
                    estado_simbolo = 'Domingo'
            elif estado == 'JUSTIFICADO':
                estado_simbolo = nombre_tipo or ''
            else:
                estado_simbolo = ''

            ws.append([
                fecha.strftime('%d/%m/%Y'),
                empleado.sucursal.nombre,
                empleado.tipo_nomina,
                empleado.codigo,
                empleado.nombre,
                empleado.departamento,
                entrada,
                salida,
                estado_rh_display or '',
                estado_simbolo,
            ])

            fila_actual = ws.max_row
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=fila_actual, column=10).fill = fill


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_sucursal = empleados[0].sucursal.nombre if empleados else 'sucursal'
    if usar_rango:
        filename = f"asistencias_{nombre_sucursal}_{fecha_inicio.strftime('%d-%m')}_al_{fecha_final.strftime('%d-%m-%Y')}.xlsx"
    else:
        filename = f"asistencias_{nombre_sucursal}_{fecha_inicio.strftime('%d-%m-%Y')}.xlsx"

    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
@grupo_requerido('rrhh')
def generar_enlace_reset_view(request, username):
    try:
        user = User.objects.get(username=username)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        token = default_token_generator.make_token(user)

        # Detecta el host autom√°ticamente desde la solicitud
        dominio = request.get_host()  # ejemplo: 192.168.11.12:8005
        enlace = f"http://{dominio}:8005/marcaje/reset/{uid}/{token}/"

        return JsonResponse({'enlace': enlace})
    except User.DoesNotExist:
        return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
